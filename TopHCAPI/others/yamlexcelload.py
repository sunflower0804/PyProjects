import os
import yaml
from openpyxl import load_workbook



def loadyaml(filename):
    #首先拿到yaml文件
    files = open(filename, 'r', encoding='utf-8')
    #然后读取yaml文件中的数据
    data = yaml.load(files, Loader=yaml.FullLoader)
    return data

class ExcelReader:
    def __init__(self, excel_file, sheet=None, title=True):
        self._file = excel_file
        if not os.path.exists(self._file):
            raise FileNotFoundError(f'execl file not found: {self._file}')
        if not self._file.endswith('.xlsx'):
            raise TypeError(f'only support .xlsx type excel file')
        excel = load_workbook(self._file, data_only=True)

        if sheet is None:
            self._s = excel.active
        elif isinstance(sheet, int):
            self._s = excel.worksheets[sheet]
        elif isinstance(sheet, str):
            self._s = excel[sheet]
        else:
            raise TypeError(f'sheet page typeerror ，only support int/str/None，not {type(sheet)} ')
        self._title = title
        self._data = None

    @property
    def data(self):
        if not self._data:
            if not self._title:
                self._data = [[cell.value for cell in self._s[row]] for row in range(1, self._s.max_row + 1)]
            else:
                title = [cell.value for cell in self._s[1]]
                self._data = [dict(zip(title, [cell.value for cell in self._s[row]])) for row in
                              range(2, self._s.max_row + 1)]
        return self._data
